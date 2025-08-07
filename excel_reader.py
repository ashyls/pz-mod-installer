from openpyxl import load_workbook

def get_excel_columns(file_path='Book1.xlsx', desired_columns=['D', 'E', 'G']):
    """
    Get specific columns from an Excel file and return them as a dictionary.
    
    Args:
        file_path (str): Path to the Excel file
        desired_columns (list): List of column letters to extract (e.g., ['B', 'D', 'F'])
    
    Returns:
        dict: Dictionary with column letters as keys and column data as values
    """
    try:
        wb = load_workbook(filename=file_path)
        ws = wb.active
        
        result = {}
        
        for col_letter in desired_columns:
            column_data = []
            for cell in ws[col_letter]:
                column_data.append(cell.value)
            
            while column_data and column_data[-1] is None:
                column_data.pop()
                
            result[f'Column_{col_letter}'] = column_data
        
        return result
    
    except FileNotFoundError:
        return {'error': f'File {file_path} not found'}
    except Exception as e:
        return {'error': f'An error occurred: {str(e)}'}